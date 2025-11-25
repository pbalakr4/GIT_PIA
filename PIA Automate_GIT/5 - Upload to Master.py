
#!/usr/bin/env python
"""
Enhanced version:
- Avoid duplicates when copying rows.
- For 'Master': skip if ID already exists.
- For 'Keyword to ID mapped': skip if (ID, Keyword, Category, Type of Identifier) exists.
"""

from openpyxl import load_workbook

# Paths
EXTRACT_PATH = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\Documents\PIA Automate\Extract.xlsx"
MASTER_PATH = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\T-Ads - Privacy & CyberSecurity\Privacy\PIA Files\T-Ads PIAs Automation\Master.xlsx"

# Sheet names
SRC_SHEET_1 = "Raw Extract"
DST_SHEET_1 = "Master"

SRC_SHEET_2 = "ID to PD Mapping"
DST_SHEET_2 = "Keyword to ID mapped"


def get_existing_ids(ws):
    """Return a set of IDs from column 1 of the given worksheet."""
    return {str(row[0].value).strip() for row in ws.iter_rows(min_row=2, max_col=1) if row[0].value}


def get_existing_combinations(ws):
    """Return a set of (ID, Keyword, Category, Type) tuples from the given worksheet."""
    combos = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[2] and row[3]:  # Ensure all fields exist
            combos.add((str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip()))
    return combos


def copy_rows_master(src_ws, dst_ws, existing_ids):
    """Copy rows to Master sheet, skipping duplicates based on ID."""
    rows_copied = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip() not in existing_ids:  # Check ID
            dst_ws.append(row)
            existing_ids.add(str(row[0]).strip())  # Update set
            rows_copied += 1
    return rows_copied


def copy_rows_keyword_mapping(src_ws, dst_ws, existing_combos):
    """Copy rows to Keyword sheet, skipping duplicates based on (ID, Keyword, Category, Type)."""
    rows_copied = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[2] and row[3]:
            combo = (str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip())
            if combo not in existing_combos:
                dst_ws.append(row)
                existing_combos.add(combo)
                rows_copied += 1
    return rows_copied


def main():
    # Load workbooks
    src_wb = load_workbook(EXTRACT_PATH, data_only=True)
    dst_wb = load_workbook(MASTER_PATH)

    # Validate sheets
    if SRC_SHEET_1 not in src_wb.sheetnames or SRC_SHEET_2 not in src_wb.sheetnames:
        raise ValueError("Source sheets missing in Extract.xlsx")
    if DST_SHEET_1 not in dst_wb.sheetnames or DST_SHEET_2 not in dst_wb.sheetnames:
        raise ValueError("Destination sheets missing in Master.xlsx")

    src_ws1 = src_wb[SRC_SHEET_1]
    src_ws2 = src_wb[SRC_SHEET_2]
    dst_ws1 = dst_wb[DST_SHEET_1]
    dst_ws2 = dst_wb[DST_SHEET_2]

    # Get existing data
    existing_ids = get_existing_ids(dst_ws1)
    existing_combos = get_existing_combinations(dst_ws2)

    # Copy rows with duplicate checks
    rows_copied_1 = copy_rows_master(src_ws1, dst_ws1, existing_ids)
    rows_copied_2 = copy_rows_keyword_mapping(src_ws2, dst_ws2, existing_combos)

    # Save destination workbook
    dst_wb.save(MASTER_PATH)

    # Print summary
    print(f"Rows copied to '{DST_SHEET_1}': {rows_copied_1}")
    print(f"Rows copied to '{DST_SHEET_2}': {rows_copied_2}")


if __name__ == "__main__":
    main()
