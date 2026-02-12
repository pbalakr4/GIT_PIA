
import os
import re
import sys
from typing import Dict, List, Optional, Tuple

# 3rd-party libraries required:
#   pip install openpyxl PyPDF2
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyPDF2 import PdfReader


# ==============================================================================
# Configuration (paths provided by you)
# ==============================================================================
PDF_FOLDER = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\Documents\PIA Automate\Consolidatedpdfs"
EXTRACT_PATH = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\Documents\PIA Automate\Extract.xlsx"

RAW_SHEET_NAME = "Raw Extract"
FULL_SHEET_NAME = "full pdf content"

# Columns A..G to copy (7 columns)
NUM_COLS_COPY = 7  # A..G
PDFCONTENT_COL_IDX = 8  # H

# Separator inserted before section-like headings (1, 1.1, 1.12, ...)
SECTION_SEPARATOR = "-" * 80


# ==============================================================================
# Helpers
# ==============================================================================

def extract_id_from_filename(filename: str) -> Optional[str]:
    """
    Extract ID from a filename in the format: <anything>_<digits>.pdf
    Returns the ID as a string (digits) or None if not found.
    """
    name, ext = os.path.splitext(filename)
    if ext.lower() != ".pdf":
        return None
    # Extract digits after the last underscore at the end of the basename
    m = re.search(r"_(\d+)$", name)
    if not m:
        return None
    return m.group(1)


def load_or_create_workbook(path: str) -> Workbook:
    """Load an existing workbook or raise a clear error if not found."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found at: {path}")
    return load_workbook(path)


def get_or_create_sheet(wb: Workbook, name: str) -> Worksheet:
    """Get an existing sheet or create a new one if it doesn't exist."""
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)


def read_headers(ws: Worksheet, num_cols: int) -> List[Optional[str]]:
    """Read headers (Row 1) for the specified number of columns."""
    return [ws.cell(row=1, column=col).value for col in range(1, num_cols + 1)]


def ensure_headers(
    ws_full: Worksheet,
    headers_from_raw: List[Optional[str]],
    pdfcontent_header: str = "pdfcontent",
) -> None:
    """
    Ensure that row 1 in 'full pdf content' contains:
      - Columns A..G: headers copied from Raw Extract
      - Column H: 'pdfcontent'
    If headers are missing or different, write/update them.
    """
    for idx, hdr in enumerate(headers_from_raw, start=1):
        ws_full.cell(row=1, column=idx, value=str(hdr) if hdr is not None else "")
    ws_full.cell(row=1, column=PDFCONTENT_COL_IDX, value=pdfcontent_header)


def gather_existing_ids(ws_full: Worksheet) -> set:
    """Return a set of IDs already present in Column A (excluding header)."""
    existing = set()
    for row in ws_full.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        cell_val = row[0]
        if cell_val is None:
            continue
        existing.add(str(cell_val).strip())
    return existing


def build_raw_extract_index(ws_raw: Worksheet, num_cols: int) -> Dict[str, Tuple]:
    """
    Build an index from ID => row values (A..G) from Raw Extract.
    Row 1 is assumed to be headers and skipped.
    """
    index = {}
    for r in range(2, ws_raw.max_row + 1):
        id_cell = ws_raw.cell(row=r, column=1).value  # Column A
        if id_cell is None:
            continue
        str_id = str(id_cell).strip()
        values = tuple(ws_raw.cell(row=r, column=c).value for c in range(1, num_cols + 1))
        index[str_id] = values
    return index


def get_next_empty_row(ws: Worksheet) -> int:
    """Get the next empty row (appends after the last non-empty row)."""
    return ws.max_row + 1 if ws.max_row else 1


# ==============================================================================
# PDF text processing (with duplicate-line cleanup)
# ==============================================================================

def _normalize_line_for_compare(line: str) -> str:
    """
    Normalize a line for duplicate detection:
      - collapse multiple spaces/tabs
      - strip leading/trailing whitespace
      - lowercase
    """
    return re.sub(r"[ \t]+", " ", line).strip().lower()


def _deduplicate_nearby_lines(raw_text: str) -> str:
    """
    Remove duplicated lines that often occur when PDF extraction reads the same
    text twice (e.g., table structures or two columns yielding duplicated labels).

    Rules:
      - If a non-empty line is identical (after normalization) to the last
        non-empty line we've *kept*, drop it. This also covers cases where
        duplicates are separated only by blank lines.
      - Collapse runs of multiple blank lines to a single blank line.
    """
    lines = raw_text.splitlines()
    out: List[str] = []
    last_non_empty_norm: Optional[str] = None

    for line in lines:
        normalized = _normalize_line_for_compare(line)

        if normalized == "":
            # Collapse multiple blank lines
            if len(out) == 0 or out[-1].strip() != "":
                out.append("")  # keep a single blank line
            continue

        # Skip if this non-empty line duplicates the last non-empty kept line
        if last_non_empty_norm is not None and normalized == last_non_empty_norm:
            continue

        # Keep this line
        out.append(line.strip())
        last_non_empty_norm = normalized

    # Trim leading/trailing blank lines
    while out and out[0].strip() == "":
        out.pop(0)
    while out and out[-1].strip() == "":
        out.pop()

    return "\n".join(out)


def _insert_section_separators(text: str) -> str:
    """
    Add a separator line BEFORE lines that *look* like section headers:
    e.g., "1 Title", "1.1 Subtitle", "1.12 Another" at the start of a line.
    """
    formatted = re.sub(
        r"(?m)^(?=\d+(?:\.\d+)*\s+)",
        f"\n{SECTION_SEPARATOR}\n",
        text,
    )
    if not formatted.startswith(SECTION_SEPARATOR):
        formatted = f"{SECTION_SEPARATOR}\n" + formatted
    return formatted


def extract_pdf_text(pdf_path: str) -> str:
    """
    Extract text from a PDF file using PyPDF2 and clean it up:
      1) join pages with markers
      2) normalize whitespace
      3) de-duplicate nearby identical lines (fixes 'ID', 'Name', etc. repeated)
      4) insert section separators for readability
    """
    text_parts: List[str] = []

    with open(pdf_path, "rb") as f:
        reader = PdfReader(f)
        for i, page in enumerate(reader.pages, start=1):
            page_text = page.extract_text() or ""
            text_parts.append(f"\n\n--- Page {i} ---\n")
            text_parts.append(page_text)

    raw_text = "".join(text_parts)

    # Normalize intra-line whitespace and excessive newlines
    raw_text = re.sub(r"[ \t]+", " ", raw_text)
    raw_text = re.sub(r"\n{3,}", "\n\n", raw_text)

    # Key fix: remove duplicate lines that appear twice due to PDF layout
    deduped = _deduplicate_nearby_lines(raw_text)

    # Add separators before section-like headings
    formatted = _insert_section_separators(deduped)

    return formatted.strip()


# ==============================================================================
# Excel writing
# ==============================================================================

def append_row_with_pdf_content(
    ws_full: Worksheet,
    values_a_to_g: Tuple,
    pdf_text: str
) -> None:
    """
    Append a new row to 'full pdf content' with A..G values and H as pdfcontent.
    """
    next_row = get_next_empty_row(ws_full)
    # Write A..G
    for idx, val in enumerate(values_a_to_g, start=1):
        ws_full.cell(row=next_row, column=idx, value=val)
    # Write H (pdfcontent)
    ws_full.cell(row=next_row, column=PDFCONTENT_COL_IDX, value=pdf_text)


# ==============================================================================
# Main
# ==============================================================================

def main():
    # Pre-flight checks
    if not os.path.isdir(PDF_FOLDER):
        print(f"[ERROR] PDF folder not found: {PDF_FOLDER}")
        sys.exit(1)
    if not os.path.exists(EXTRACT_PATH):
        print(f"[ERROR] Excel file not found: {EXTRACT_PATH}")
        sys.exit(1)

    # Load workbook and sheets
    wb = load_or_create_workbook(EXTRACT_PATH)

    if RAW_SHEET_NAME not in wb.sheetnames:
        print(f"[ERROR] Sheet '{RAW_SHEET_NAME}' not found in workbook.")
        sys.exit(1)

    ws_raw = wb[RAW_SHEET_NAME]
    ws_full = get_or_create_sheet(wb, FULL_SHEET_NAME)

    # Read headers from Raw Extract (A..G) and ensure in full pdf content (A..G + H)
    raw_headers = read_headers(ws_raw, NUM_COLS_COPY)
    ensure_headers(ws_full, raw_headers, pdfcontent_header="pdfcontent")

    # Build index of Raw Extract: ID (col A) -> tuple(A..G)
    raw_index = build_raw_extract_index(ws_raw, NUM_COLS_COPY)

    # Gather existing IDs already present in full pdf content's Column A
    existing_ids = gather_existing_ids(ws_full)

    print(f"[INFO] Raw Extract rows indexed: {len(raw_index)}")
    print(f"[INFO] Existing IDs in '{FULL_SHEET_NAME}': {len(existing_ids)}")

    # Iterate over PDF files
    pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith(".pdf")]
    pdf_files.sort()

    added_count = 0
    skipped_no_id = 0
    skipped_not_in_raw = 0
    skipped_duplicate = 0
    errors = 0

    for filename in pdf_files:
        pdf_path = os.path.join(PDF_FOLDER, filename)
        file_id = extract_id_from_filename(filename)

        if not file_id:
            print(f"[WARN] Skipping (no ID pattern) => {filename}")
            skipped_no_id += 1
            continue

        # ID match check against Raw Extract
        if file_id not in raw_index:
            print(f"[WARN] Skipping (ID not found in Raw Extract) => {filename} [ID={file_id}]")
            skipped_not_in_raw += 1
            continue

        # Duplicate check against full pdf content Column A
        if file_id in existing_ids:
            print(f"[INFO] Skipping duplicate => {filename} [ID={file_id}]")
            skipped_duplicate += 1
            continue

        # Extract PDF text
        try:
            pdf_text = extract_pdf_text(pdf_path)
        except Exception as e:
            print(f"[ERROR] Failed to read PDF: {filename} ({e})")
            errors += 1
            continue

        # Append row: A..G from Raw Extract + H = pdfcontent
        try:
            values_a_to_g = raw_index[file_id]
            append_row_with_pdf_content(ws_full, values_a_to_g, pdf_text)
            existing_ids.add(file_id)  # update set to avoid duplicates in same run
            added_count += 1
            print(f"[OK] Added ID={file_id} from {filename}")
        except Exception as e:
            print(f"[ERROR] Failed to write row for {filename}: {e}")
            errors += 1

    # Save workbook
    try:
        wb.save(EXTRACT_PATH)
        print(f"\n[SAVED] Workbook updated: {EXTRACT_PATH}")
    except PermissionError:
        print("\n[ERROR] Could not save workbook (is it open in Excel?). Please close it and re-run.")
    except Exception as e:
        print(f"\n[ERROR] Failed to save workbook: {e}")

    # Summary
    print("\n----- Summary -----")
    print(f"PDFs scanned           : {len(pdf_files)}")
    print(f"Rows added             : {added_count}")
    print(f"Skipped (no ID)        : {skipped_no_id}")
    print(f"Skipped (ID not in Raw): {skipped_not_in_raw}")
    print(f"Skipped (duplicate)    : {skipped_duplicate}")
    print(f"Errors                 : {errors}")


if __name__ == "__main__":
    # IMPORTANT:
    # - Close the Excel file (Extract.xlsx) before running this script, or saving will fail.
    # - Ensure you have `openpyxl` and `PyPDF2` installed:
    #     pip install openpyxl PyPDF2
    main()
