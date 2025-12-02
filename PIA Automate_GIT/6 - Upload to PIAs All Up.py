
import os
import shutil
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# =========================
# Configurations
# =========================
SOURCE_DIR = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\Documents\PIA Automate\Consolidatedpdfs"
DEST_DIR = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\T-Ads - Privacy & CyberSecurity\Privacy\PIA Files\T-Ads PIAs Automation\PIAs All Up"
MASTER_PATH = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\T-Ads - Privacy & CyberSecurity\Privacy\PIA Files\T-Ads PIAs Automation\Master.xlsx"
MASTER_SHEET_NAME = "Master"
LINK_SHEET_NAME = "PIAs Link"
COPY_RECURSIVE = True

# SharePoint Base URL
SHAREPOINT_BASE_URL = (
    "https://tmobileusa.sharepoint.com/:b:/r/sites/TMS20/Shared%20Documents/"
    "Product/Privacy%20%26%20CyberSecurity/Privacy/PIA%20Files/"
    "T-Ads%20PIAs%20Automation/PIAs%20All%20Up"
)

# =========================
# Utility Functions
# =========================
def normalize_cell_value(val):
    return str(val).strip() if val else ""

def last_data_row_in_col_a(ws: Worksheet) -> int:
    for row in range(ws.max_row, 1, -1):
        if normalize_cell_value(ws.cell(row=row, column=1).value):
            return row
    return 1

def get_existing_ids_strict(ws: Worksheet) -> set:
    ids = set()
    for row in range(2, ws.max_row + 1):
        val = normalize_cell_value(ws.cell(row=row, column=1).value)
        if val:
            ids.add(val.lower())
    return ids

# =========================
# Part 1: Copy PDFs
# =========================
def copy_pdfs(source_dir: str, dest_dir: str, recursive: bool = True) -> Tuple[int, int, int]:
    """
    Copies all .pdf files from source_dir to dest_dir.
    Returns (new_files_count, overwritten_files_count, total_files_in_dest).
    """
    if not os.path.exists(dest_dir):
        os.makedirs(dest_dir)

    new_files_count = 0
    overwritten_files_count = 0

    for root, dirs, files in os.walk(source_dir):
        for file in files:
            if file.lower().endswith(".pdf"):
                src_path = os.path.join(root, file)
                dest_path = os.path.join(dest_dir, file)

                if os.path.exists(dest_path):
                    # File exists → overwrite
                    shutil.copy2(src_path, dest_path)
                    overwritten_files_count += 1
                else:
                    # File is new → copy
                    shutil.copy2(src_path, dest_path)
                    new_files_count += 1

        if not recursive:
            break

    # Total files in destination after operation
    total_files_in_dest = len([f for f in os.listdir(dest_dir) if f.lower().endswith(".pdf")])

    print(f"[INFO] New PDFs copied: {new_files_count}")
    print(f"[INFO] PDFs overwritten: {overwritten_files_count}")
    print(f"[INFO] Total PDFs in destination folder: {total_files_in_dest}")

    return new_files_count, overwritten_files_count, total_files_in_dest

# =========================
# Part 2: Excel Processing
# =========================
def collect_rows_from_master(master_sheet: Worksheet) -> List[Tuple[str, str]]:
    rows: List[Tuple[str, str]] = []
    for row_idx in range(2, master_sheet.max_row + 1):
        a_val = normalize_cell_value(master_sheet.cell(row=row_idx, column=1).value)
        b_val = normalize_cell_value(master_sheet.cell(row=row_idx, column=2).value)
        if a_val and b_val:
            rows.append((a_val, b_val))
    return rows

def ensure_sheet(workbook, sheet_name: str) -> Worksheet:
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    else:
        return workbook.create_sheet(title=sheet_name)

def ensure_headers(link_ws: Worksheet, master_ws: Worksheet) -> None:
    col1_header = normalize_cell_value(link_ws.cell(row=1, column=1).value)
    col2_header = normalize_cell_value(link_ws.cell(row=1, column=2).value)
    master_col1_header = normalize_cell_value(master_ws.cell(row=1, column=1).value) or "Column 1"
    master_col2_header = normalize_cell_value(master_ws.cell(row=1, column=2).value) or "Column 2"
    if not col1_header:
        link_ws.cell(row=1, column=1, value=master_col1_header)
    if not col2_header:
        link_ws.cell(row=1, column=2, value=master_col2_header)

    if normalize_cell_value(link_ws.cell(row=1, column=3).value) != "Description":
        link_ws.cell(row=1, column=3, value="Description")
    if normalize_cell_value(link_ws.cell(row=1, column=4).value) != "hyperlink address":
        link_ws.cell(row=1, column=4, value="hyperlink address")

def copy_unique_rows(master_ws: Worksheet, link_ws: Worksheet) -> int:
    master_rows = collect_rows_from_master(master_ws)
    existing_ids = get_existing_ids_strict(link_ws)
    start_row = last_data_row_in_col_a(link_ws) + 1
    appended = 0

    for a_val, b_val in master_rows:
        key = a_val.lower()
        if key in existing_ids:
            continue
        link_ws.cell(row=start_row, column=1, value=a_val)
        link_ws.cell(row=start_row, column=2, value=b_val)
        start_row += 1
        appended += 1
        existing_ids.add(key)

    print(f"[INFO] Unique rows appended to '{LINK_SHEET_NAME}': {appended}")
    return appended

def update_description_and_links(master_ws: Worksheet, link_ws: Worksheet, dest_dir: str) -> int:
    last_row = last_data_row_in_col_a(link_ws)
    count = 0

    for row_idx in range(2, last_row + 1):
        a_val = normalize_cell_value(link_ws.cell(row=row_idx, column=1).value)
        b_val = normalize_cell_value(link_ws.cell(row=row_idx, column=2).value)
        if not (a_val and b_val):
            continue

        # Column 3: Description from Master (Column J)
        description = normalize_cell_value(master_ws.cell(row=row_idx, column=10).value)
        link_ws.cell(row=row_idx, column=3, value=description)

        # Column 4: hyperlink address using SharePoint URL
        file_name = f"{b_val}_{a_val}.pdf"
        address = f"{SHAREPOINT_BASE_URL}/{file_name}"
        link_ws.cell(row=row_idx, column=4, value=address)

        count += 1

    print(f"[INFO] Updated Description and hyperlink address: {count}")
    return count

def process_master_excel(master_path: str, dest_dir: str) -> None:
    if not os.path.exists(master_path):
        raise FileNotFoundError(f"Master Excel not found: {master_path}")

    wb = load_workbook(master_path)
    if MASTER_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{MASTER_SHEET_NAME}' not found in {master_path}")

    master_ws = wb[MASTER_SHEET_NAME]
    link_ws = ensure_sheet(wb, LINK_SHEET_NAME)

    ensure_headers(link_ws, master_ws)
    copy_unique_rows(master_ws, link_ws)
    update_description_and_links(master_ws, link_ws, dest_dir)

    wb.save(master_path)
    print(f"[INFO] Saved changes to: {master_path}")

# =========================
# Main Entry
# =========================
if __name__ == "__main__":
    print("[INFO] Starting Part 1: Copy PDFs...")
    new_count, overwritten_count, total_count = copy_pdfs(SOURCE_DIR, DEST_DIR, recursive=COPY_RECURSIVE)

    print("[INFO] Starting Part 2: Update Excel and Hyperlinks...")
    process_master_excel(MASTER_PATH, DEST_DIR)

    print("[INFO] Completed all tasks successfully.")
    print(f"       New PDFs copied: {new_count}")
    print(f"       PDFs overwritten: {overwritten_count}")
    print(f"       Total PDFs in folder: {total_count}")
