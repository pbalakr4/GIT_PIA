
#!/usr/bin/env python
from openpyxl import load_workbook

EXTRACT_PATH = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\Documents\PIA Automate\Extract.xlsx"
MASTER_PATH = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\T-Ads - Privacy & CyberSecurity\Privacy\PIA Files\T-Ads PIAs Automation\Master.xlsx"

SRC_SHEET_1 = "Raw Extract"
DST_SHEET_1 = "Master"

SRC_SHEET_2 = "ID to PD Mapping"
DST_SHEET_2 = "Keyword to ID mapped"

COMPOSITE_HEADERS = ["ID", "Keywords", "Category", "Type of Identifier"]

def get_header_map(ws):
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    return {h.lower(): idx for idx, h in enumerate(headers)}

def normalize(val):
    return str(val).strip().lower() if val else ""

def get_existing_ids(ws):
    ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            ids.add(normalize(row[0]))
    return ids

def get_existing_combos(ws, header_map):
    indices = [header_map[h.lower()] for h in COMPOSITE_HEADERS if h.lower() in header_map]
    combos = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        values = [normalize(row[idx]) if idx < len(row) else "" for idx in indices]
        combos.add(tuple(values))
    return combos

def copy_rows_master(src_ws, dst_ws, existing_ids):
    rows_copied = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        id_val = normalize(row[0])
        if id_val and id_val not in existing_ids:
            dst_ws.append(row)
            existing_ids.add(id_val)
            rows_copied += 1
            print(f"[COPIED Master] ID={id_val}")
        else:
            print(f"[SKIPPED Master] ID={id_val} (duplicate)")
    return rows_copied

def copy_rows_keyword_mapping(src_ws, dst_ws, existing_combos, header_map):
    rows_copied = 0
    skipped_due_to_duplicate = 0
    indices = [header_map[h.lower()] for h in COMPOSITE_HEADERS if h.lower() in header_map]
    for row_num, row in enumerate(src_ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            print(f"[SKIPPED Row {row_num}] Empty row")
            continue
        values = [normalize(row[idx]) if idx < len(row) else "" for idx in indices]
        combo = tuple(values)
        if combo not in existing_combos:
            dst_ws.append(row)
            existing_combos.add(combo)
            rows_copied += 1
            print(f"[COPIED Mapping] Row {row_num} Composite Key={combo}")
        else:
            skipped_due_to_duplicate += 1
            print(f"[SKIPPED Mapping] Row {row_num} Duplicate Composite Key={combo}")
    print(f"Skipped {skipped_due_to_duplicate} rows due to duplicate composite keys.")
    return rows_copied

def main():
    src_wb = load_workbook(EXTRACT_PATH, data_only=True)
    dst_wb = load_workbook(MASTER_PATH)

    src_ws1 = src_wb[SRC_SHEET_1]
    src_ws2 = src_wb[SRC_SHEET_2]
    dst_ws1 = dst_wb[DST_SHEET_1]
    dst_ws2 = dst_wb[DST_SHEET_2]

    header_map_dst2 = get_header_map(dst_ws2)

    existing_ids = get_existing_ids(dst_ws1)
    existing_combos = get_existing_combos(dst_ws2, header_map_dst2)

    print("\n--- Copying Raw Extract → Master ---")
    rows_copied_1 = copy_rows_master(src_ws1, dst_ws1, existing_ids)

    print("\n--- Copying ID to PD Mapping → Keyword to ID mapped ---")
    rows_copied_2 = copy_rows_keyword_mapping(src_ws2, dst_ws2, existing_combos, header_map_dst2)

    dst_wb.save(MASTER_PATH)

    print("\n--- Summary ---")
    print(f"Rows copied to '{DST_SHEET_1}': {rows_copied_1}")
    print(f"Rows copied to '{DST_SHEET_2}': {rows_copied_2}")

if __name__ == "__main__":
    main()
