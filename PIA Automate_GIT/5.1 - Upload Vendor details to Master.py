
#!/usr/bin/env python
from openpyxl import load_workbook

EXTRACT_PATH = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\Documents\PIA Automate\Extract.xlsx"
MASTER_PATH = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\T-Ads - Privacy & CyberSecurity\Privacy\PIA Files\T-Ads PIAs Automation\Master.xlsx"

VENDOR_SRC_SHEET = "Vendor Extraction"
VENDOR_DST_SHEET = "Vendor Details"


def copy_vendor_details():
    # Load workbooks
    try:
        src_wb = load_workbook(EXTRACT_PATH, data_only=True)
    except Exception as e:
        print(f"[ERROR] Unable to open Extract file: {EXTRACT_PATH}\n{e}")
        return

    try:
        dst_wb = load_workbook(MASTER_PATH)
    except Exception as e:
        print(f"[ERROR] Unable to open Master file: {MASTER_PATH}\n{e}")
        return

    # Validate source sheet exists
    if VENDOR_SRC_SHEET not in src_wb.sheetnames:
        print(f"[ERROR] Source sheet '{VENDOR_SRC_SHEET}' not found in Extract file.")
        return

    src_ws = src_wb[VENDOR_SRC_SHEET]

    # Prepare destination sheet
    if VENDOR_DST_SHEET in dst_wb.sheetnames:
        dst_ws = dst_wb[VENDOR_DST_SHEET]
        # If there are any rows beyond the header, delete them (preserve headers in row 1)
        if dst_ws.max_row >= 2:
            dst_ws.delete_rows(2, dst_ws.max_row - 1)
    else:
        dst_ws = dst_wb.create_sheet(title=VENDOR_DST_SHEET)
        # If we create a new sheet, write the header from the source
        src_header = [cell.value for cell in next(src_ws.iter_rows(min_row=1, max_row=1))]
        dst_ws.append(src_header)

    # Read source header to support consistent data copy
    src_header = [cell.value for cell in next(src_ws.iter_rows(min_row=1, max_row=1))]

    # Copy only data rows (starting from row 2 in source)
    rows_copied = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        dst_ws.append(row)
        rows_copied += 1

    # Save changes
    try:
        dst_wb.save(MASTER_PATH)
    except Exception as e:
        print(f"[ERROR] Unable to save Master file: {MASTER_PATH}\n{e}")
        return

    print(f"[SUCCESS] Cleared existing rows from 2 onward and copied {rows_copied} data rows "
          f"from '{VENDOR_SRC_SHEET}' to '{VENDOR_DST_SHEET}'. Headers preserved.")


def main():
    copy_vendor_details()


if __name__ == "__main__":
    main()
