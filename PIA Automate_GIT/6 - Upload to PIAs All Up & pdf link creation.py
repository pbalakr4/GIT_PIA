

import os
import re
import shutil
from typing import List, Tuple, Dict, Optional
from urllib.parse import quote
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# =========================
# Configurations
# =========================
SOURCE_DIR = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\Documents\PIA Automate\Consolidatedpdfs"
DEST_DIR = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\T-Ads - Privacy & CyberSecurity\Privacy\PIA Files\T-Ads PIAs Automation\PIAs All Up"
MASTER_PATH = r"C:\Users\PBalakr4\OneDrive - T-Mobile USA\T-Ads - Privacy & CyberSecurity\Privacy\PIA Files\T-Ads PIAs Automation\Master.xlsx"
MASTER_SHEET_NAME = "Master"
LINK_SHEET_NAME = "PIAs Link"
COPY_RECURSIVE = True

# =========================
# SharePoint Base URL
# =========================
# IMPORTANT: Use your actual SharePoint folder URL (no trailing slash).
# Keep spaces in the base path pre-encoded (e.g., Shared%20Documents).
SHAREPOINT_BASE_URL = (
    "https://tmobileusa.sharepoint.com/:b:/r/sites/TMS20/Shared%20Documents/"
    "Product/Privacy%20%26%20CyberSecurity/Privacy/PIA%20Files/"
    "T-Ads%20PIAs%20Automation/PIAs%20All%20Up"
)

# Optional query suffix to open in web viewer
# Examples: "", "?web=1", "?csf=1&web=1", "?csf=1&web=1&e=cTGxUz"
QUERY_SUFFIX = "?csf=1&web=1"

# =========================
# Utility Functions
# =========================
def normalize_cell_value(val) -> str:
    """Trimmed string for general use (internal spaces preserved)."""
    return str(val).strip() if val else ""

def last_data_row_in_col_a(ws: Worksheet) -> int:
    """Find the last non-empty row in column A."""
    for row in range(ws.max_row, 1, -1):
        if normalize_cell_value(ws.cell(row=row, column=1).value):
            return row
    return 1

def get_existing_ids_strict(ws: Worksheet) -> set:
    """Collect existing IDs (column A) in a case-insensitive set."""
    ids = set()
    for row in range(2, ws.max_row + 1):
        val = normalize_cell_value(ws.cell(row=row, column=1).value)
        if val:
            ids.add(val.lower())
    return ids

def encode_filename_for_url(filename: str) -> str:
    """
    Encode ONLY the filename for inclusion in a URL path segment.
    - Preserves '-', '_', '.', '(' and ')'
    - Encodes spaces (%20) and '&' (%26)
    """
    return quote(filename, safe='-_.()')

def build_sharepoint_file_url(base_url: str, filename: str) -> str:
    """Base URL + encoded filename + optional query suffix."""
    base = base_url.strip().rstrip('/')
    encoded_filename = encode_filename_for_url(filename)
    url = f"{base}/{encoded_filename}"
    if QUERY_SUFFIX:
        url += QUERY_SUFFIX if QUERY_SUFFIX.startswith('?') else f"?{QUERY_SUFFIX}"
    return url

# =========================
# ID ↔ Filename mapping from DEST_DIR
# =========================
# Matches ..._{ID}.pdf OR ...  _{ID}.pdf (optional space before underscore), case-insensitive
ID_SUFFIX_PATTERN = re.compile(r"(?:\s)?_(\d+)\.pdf$", re.IGNORECASE)

def scan_dest_dir_for_id_map(dest_dir: str) -> Dict[str, List[str]]:
    """
    Scan dest_dir for PDFs; build a map: id_lower -> [filenames].
    Matches filenames that end with ' _{ID}.pdf' or '_{ID}.pdf', case-insensitive.
    """
    id_map: Dict[str, List[str]] = {}
    if not os.path.isdir(dest_dir):
        print(f"[WARN] DEST_DIR not found: {dest_dir}", flush=True)
        return id_map

    for f in os.listdir(dest_dir):
        fl = f.lower()
        if not fl.endswith(".pdf"):
            continue
        m = ID_SUFFIX_PATTERN.search(f)
        if m:
            id_val = m.group(1).lower()
            id_map.setdefault(id_val, []).append(f)

    print(f"[INFO] Scanned DEST_DIR: found {sum(len(v) for v in id_map.values())} PDF(s) with ID suffix", flush=True)
    return id_map

def choose_best_filename(candidates: List[str], name_hint: Optional[str]) -> str:
    """
    If multiple files share the same ID, pick the one whose prefix best matches name_hint (Column B),
    otherwise return a deterministic first (sorted).
    """
    if not candidates:
        raise ValueError("No candidates to choose from.")
    if len(candidates) == 1 or not name_hint:
        return sorted(candidates)[0]

    def norm(s: str) -> str:
        return s.lower().replace('/', '_').replace('\\', '_').strip()

    hint = norm(name_hint)
    best = None
    best_score = -1
    for f in candidates:
        # prefix before the last underscore + ID (allow for optional space)
        fl = f.lower()
        pos = fl.rfind("_")
        prefix = f[:pos].rstrip()
        pf = norm(prefix)
        score = 0
        if pf == hint:
            score = 3
        elif pf.startswith(hint) or hint.startswith(pf):
            score = 2
        elif hint in pf or pf in hint:
            score = 1
        if score > best_score:
            best = f
            best_score = score
    return best or sorted(candidates)[0]

# =========================
# Part 1: Copy PDFs
# =========================
def copy_pdfs(source_dir: str, dest_dir: str, recursive: bool = True) -> Tuple[int, int, int]:
    """
    Copies all .pdf files from source_dir to dest_dir.
    Returns (new_files_count, overwritten_files_count, total_files_in_dest).
    """
    if not os.path.exists(dest_dir):
        os.makedirs(dest_dir)

    new_files_count = 0
    overwritten_files_count = 0

    for root, dirs, files in os.walk(source_dir):
        for file in files:
            if file.lower().endswith(".pdf"):
                src_path = os.path.join(root, file)
                dest_path = os.path.join(dest_dir, file)

                if os.path.exists(dest_path):
                    # File exists → overwrite
                    shutil.copy2(src_path, dest_path)
                    overwritten_files_count += 1
                else:
                    # File is new → copy
                    shutil.copy2(src_path, dest_path)
                    new_files_count += 1

        if not recursive:
            break

    # Total files in destination after operation
    total_files_in_dest = len([f for f in os.listdir(dest_dir) if f.lower().endswith(".pdf")])

    print(f"[INFO] New PDFs copied: {new_files_count}", flush=True)
    print(f"[INFO] PDFs overwritten: {overwritten_files_count}", flush=True)
    print(f"[INFO] Total PDFs in destination folder: {total_files_in_dest}", flush=True)

    return new_files_count, overwritten_files_count, total_files_in_dest

# =========================
# Part 2: Excel Processing
# =========================
def collect_rows_from_master(master_sheet: Worksheet) -> List[Tuple[str, str]]:
    """Collect (Column A, Column B) pairs from the Master sheet, skipping blanks."""
    rows: List[Tuple[str, str]] = []
    for row_idx in range(2, master_sheet.max_row + 1):
        a_val = normalize_cell_value(master_sheet.cell(row=row_idx, column=1).value)  # ID
        b_val = normalize_cell_value(master_sheet.cell(row=row_idx, column=2).value)  # Name/prefix (trimmed)
        if a_val and b_val:
            rows.append((a_val, b_val))
    print(f"[INFO] Master rows collected: {len(rows)}", flush=True)
    return rows

def ensure_sheet(workbook, sheet_name: str) -> Worksheet:
    """Return existing sheet or create it if missing."""
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    else:
        return workbook.create_sheet(title=sheet_name)

def ensure_headers(link_ws: Worksheet, master_ws: Worksheet) -> None:
    """Ensure the link sheet has headers in columns A–D."""
    col1_header = normalize_cell_value(link_ws.cell(row=1, column=1).value)
    col2_header = normalize_cell_value(link_ws.cell(row=1, column=2).value)
    master_col1_header = normalize_cell_value(master_ws.cell(row=1, column=1).value) or "Column 1"
    master_col2_header = normalize_cell_value(master_ws.cell(row=1, column=2).value) or "Column 2"
    if not col1_header:
        link_ws.cell(row=1, column=1, value=master_col1_header)
    if not col2_header:
        link_ws.cell(row=1, column=2, value=master_col2_header)

    if normalize_cell_value(link_ws.cell(row=1, column=3).value) != "Description":
        link_ws.cell(row=1, column=3, value="Description")
    if normalize_cell_value(link_ws.cell(row=1, column=4).value) != "hyperlink address":
        link_ws.cell(row=1, column=4, value="hyperlink address")

def copy_unique_rows(master_ws: Worksheet, link_ws: Worksheet) -> int:
    """Append unique rows from Master (based on Column A ID, case-insensitive) into the Link sheet."""
    master_rows = collect_rows_from_master(master_ws)
    existing_ids = get_existing_ids_strict(link_ws)
    start_row = last_data_row_in_col_a(link_ws) + 1
    appended = 0

    for a_val, b_val in master_rows:
        key = a_val.lower()
        if key in existing_ids:
            continue
        link_ws.cell(row=start_row, column=1, value=a_val)
        link_ws.cell(row=start_row, column=2, value=b_val)
        start_row += 1
        appended += 1
        existing_ids.add(key)

    print(f"[INFO] Unique rows appended to '{LINK_SHEET_NAME}': {appended}", flush=True)
    return appended

def update_description_and_links(master_ws: Worksheet, link_ws: Worksheet, dest_dir: str) -> int:
    """
    For each populated row in the Link sheet:
      - Column 3: Description from Master (Column J) using the same row index.
      - Column 4: Hyperlink address using SharePoint URL; filename chosen by matching ID at the end.
    """
    last_row = last_data_row_in_col_a(link_ws)
    count = 0

    # Build ID -> [filenames] map from DEST_DIR (once)
    id_map = scan_dest_dir_for_id_map(dest_dir)

    for row_idx in range(2, last_row + 1):
        a_val = normalize_cell_value(link_ws.cell(row=row_idx, column=1).value)  # ID
        b_val = normalize_cell_value(link_ws.cell(row=row_idx, column=2).value)  # Name/prefix (trimmed)
        if not (a_val and b_val):
            continue

        # Column 3: Description from Master (Column J)
        description = normalize_cell_value(master_ws.cell(row=row_idx, column=10).value)
        link_ws.cell(row=row_idx, column=3, value=description)

        # Choose filename by ID match from DEST_DIR; if multiple, prefer best match to b_val
        candidates = id_map.get(a_val.lower(), [])
        if candidates:
            file_name = choose_best_filename(candidates, b_val)
            src = "dest-match"
        else:
            # Fallback filename:
            # Do NOT force a space; add a space only if Column B in Master ends with a space
            b_raw = str(master_ws.cell(row=row_idx, column=2).value or "")
            safe_prefix = b_raw.replace('/', '_').replace('\\', '_')
            sep = " _" if b_raw.endswith(" ") else "_"
            file_name = f"{safe_prefix}{sep}{a_val}.pdf"
            src = "fallback"

        # Build SharePoint URL
        address = build_sharepoint_file_url(SHAREPOINT_BASE_URL, file_name)

        # Write hyperlink address visibly and set the cell hyperlink
        addr_cell = link_ws.cell(row=row_idx, column=4, value=address)
        addr_cell.hyperlink = address
        addr_cell.style = "Hyperlink"

        print(f"[ROW {row_idx}] ID={a_val} | source={src} | file='{file_name}' | url='{address}'", flush=True)
        count += 1

    print(f"[INFO] Updated Description and hyperlink address: {count}", flush=True)
    return count

def process_master_excel(master_path: str, dest_dir: str) -> None:
    """Main Excel processing entry."""
    if not os.path.exists(master_path):
        raise FileNotFoundError(f"Master Excel not found: {master_path}")

    wb = load_workbook(master_path)
    if MASTER_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{MASTER_SHEET_NAME}' not found in {master_path}")

    master_ws = wb[MASTER_SHEET_NAME]
    link_ws = ensure_sheet(wb, LINK_SHEET_NAME)

    ensure_headers(link_ws, master_ws)
    copy_unique_rows(master_ws, link_ws)
    update_description_and_links(master_ws, link_ws, dest_dir)

    wb.save(master_path)
    print(f"[INFO] Saved changes to: {master_path}", flush=True)

# =========================
# Main Entry
# =========================
if __name__ == "__main__":
    print("[INFO] Starting Part 1: Copy PDFs...", flush=True)
    new_count, overwritten_count, total_count = copy_pdfs(SOURCE_DIR, DEST_DIR, recursive=COPY_RECURSIVE)

    print("[INFO] Starting Part 2: Update Excel and Hyperlinks...", flush=True)
    process_master_excel(MASTER_PATH, DEST_DIR)

    print("[INFO] Completed all tasks successfully.", flush=True)
    print(f"New PDFs copied: {new_count}", flush=True)
    print(f"PDFs overwritten: {overwritten_count}", flush=True)




